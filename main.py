import openpyxl
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
from calculadora_iva import procesar_ventas, procesar_compras, mostrar_resultados, procesar_saldos_anteriores,\
    procesar_retenciones
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import Scrollbar

# Seteamos opciones del dataframe
pd.options.display.float_format = '{:.2f}'.format
pd.set_option('display.width', 200)
pd.set_option('display.max_columns', 8)


# Scrollbar archivos encontrados
def mostrar_archivos_encontrados(archivos_encontrados):
    ventana_archivos = tk.Toplevel(ventana)
    ventana_archivos.title("Archivos encontrados")
    ventana_archivos.geometry("700x300")

    scrollbar = tk.Scrollbar(ventana_archivos)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    archivos_text = tk.Text(ventana_archivos, yscrollcommand=scrollbar.set)
    archivos_text.pack(side=tk.LEFT, fill=tk.BOTH)
    scrollbar.config(command=archivos_text.yview)

    for archivo in archivos_encontrados:
        archivos_text.insert(tk.END, archivo + "\n")

    archivos_text.config(state=tk.DISABLED)


def seleccionar_archivo_saldos():
    archivo_saldos = filedialog.askopenfilename(title="Seleccionar archivo de saldos",
                                                filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    ruta_archivo_saldos_entry.delete(0, tk.END)
    ruta_archivo_saldos_entry.insert(0, archivo_saldos)


def iniciar_proceso():
    """
    Esta funcion ejecuta el proceso principal dentro de la aplicacion.

    El usuario selecciona la ubicacion donde se encuentran guardados los archivos de MCE y MCR,
    para luego calcular las posiciones de IVA del primer párrafo para cada contribuyente.

    """

    # Definimos la variable path que almacena la direccion ingresada por el usuario
    path = ruta_archivos_entry.get()
    path_retenciones = ruta_retenciones_entry.get()
    saldos_anteriores = None
    retenciones = None
    archivo_saldos_anteriores = ruta_archivo_saldos_entry.get()

    # Si la direccion ingresada por el usuario no existe
    while not os.path.exists(path):
        messagebox.showerror("Error",
                             "La ruta proporcionada no es válida. Por favor, ingrese una ubicación válida.")
        path = filedialog.askdirectory(initialdir="/",
                                       title="Seleccione la ubicación de los archivos libros IVA compras e IVA Ventas "
                                             "y Retenciones")

    # Mostrar archivos encontrados en la dirección indicada
    archivos_encontrados = os.listdir(path)
    mostrar_archivos_encontrados(archivos_encontrados)

    # Procesar archivos de ventas
    try:
        ventas = procesar_ventas(path)
    except Exception as e:
        messagebox.showerror("Error", f"Se ha producido un error al procesar los archivos de ventas: {e}")

    # Procesar archivos de compras
    try:
        compras = procesar_compras(path)
    except Exception as e:
        messagebox.showerror("Error", f"Se ha producido un error al procesar los archivos de compras: {e}")

    else:
        # Mostrar resultados
        if archivo_saldos_anteriores:
            saldos_anteriores = procesar_saldos_anteriores(archivo_saldos_anteriores)

        if path_retenciones:
            retenciones = procesar_retenciones(path_retenciones)

        resultados = mostrar_resultados(ventas, compras, saldos_anteriores, retenciones)

        # Definimos el nombre del archivo
        filename = 'WP_liquidaciones_IVA.xlsx'

        # Exportar resultados a excel
        with pd.ExcelWriter(f'{path}/{filename}') as writer:
            resultados.to_excel(writer,
                                sheet_name='WP_IVA',
                                index=False)

        workbook = openpyxl.load_workbook(f'{path}/{filename}')
        hoja = workbook['WP_IVA']  # Nombre de la hoja del DataFrame

        # Definir el color de fondo y la fuente
        fill = PatternFill(start_color="229954", end_color="229954", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)

        # Aplicar formato al encabezado
        for cell in hoja[1]:
            cell.fill = fill
            cell.font = font

        # Expandir todas las columnas
        for column in hoja.columns:
            max_length = 0
            column_values = [cell.value for cell in column]
            for value in column_values:
                try:
                    if len(str(value)) > max_length:
                        max_length = len(str(value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            column_letter = column[0].column_letter
            hoja.column_dimensions[column_letter].width = adjusted_width

        # Formato de contabilidad para las celdas con números
        number_format = '#,##0.00'

        # Recorrer las celdas y aplicar formato de contabilidad a las celdas con números
        for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=2, max_col=hoja.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format

        workbook.save(f'{path}/{filename}')

        messagebox.showinfo("Exportación exitosa",
                            f"Proceso Finalizado.\n\nLos resultados se exportaron a la ubicacion:\n{path}")


# UI Tkinter

# Creamos la ventana principal
ventana = tk.Tk()
ventana.title("APP Calculadora de IVA")
ventana.geometry("600x300")

# Título de la ventana principal
titulo_label = tk.Label(ventana, text="Calculadora de IVA", font=("Arial", 18, "bold"))
titulo_label.pack(pady=20)

# Frame de Ruta de Comprobantes
ruta_archivos_frame = tk.Frame(ventana)
ruta_archivos_frame.pack(pady=10)

ruta_archivos_label = tk.Label(ruta_archivos_frame, text="Ruta de archivos (MCE-MCR):")
ruta_archivos_label.pack(side=tk.LEFT, padx=5)

ruta_archivos_entry = tk.Entry(ruta_archivos_frame)
ruta_archivos_entry.pack(side=tk.LEFT, padx=5)


# Frame de Ruta de Retenciones
ruta_retenciones_frame = tk.Frame(ventana)
ruta_retenciones_frame.pack(pady=10)

ruta_retenciones_label = tk.Label(ruta_retenciones_frame, text="Ruta de Retenciones:")
ruta_retenciones_label.pack(side=tk.LEFT, padx=5)

ruta_retenciones_entry = tk.Entry(ruta_retenciones_frame)
ruta_retenciones_entry.pack(side=tk.LEFT, padx=5)

# Frame de Selección de archivo de saldos
archivo_saldos_frame = tk.Frame(ventana)
archivo_saldos_frame.pack(pady=10)

archivo_saldos_label = tk.Label(archivo_saldos_frame, text="Archivo de saldos:")
archivo_saldos_label.pack(side=tk.LEFT, padx=5)

ruta_archivo_saldos_entry = tk.Entry(archivo_saldos_frame)
ruta_archivo_saldos_entry.pack(side=tk.LEFT, padx=5)

seleccionar_archivo_saldos_button = tk.Button(archivo_saldos_frame, text="Seleccionar",
                                              command=seleccionar_archivo_saldos)

seleccionar_archivo_saldos_button.pack(side=tk.LEFT, padx=5)

# Boton inciar proceso
iniciar_button = tk.Button(ventana, text="Iniciar Proceso", command=iniciar_proceso)
iniciar_button.pack()

# Footer
footer_label = tk.Label(ventana, text="Developed by Gian Franco Lorenzo")
footer_label.pack(side=tk.BOTTOM, pady=10)

ventana.mainloop()
